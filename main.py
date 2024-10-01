import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import openpyxl

# Путь к файлу с именами студентов
students_filename = 'students.xlsx'

# Загрузка списка студентов из файла
students_df = pd.read_excel(students_filename)
students = students_df.iloc[:, 0].tolist()

# Создаем даты от 9 сентября до конца четверти (например, 25 декабря)
start_date = datetime.date(2024, 9, 9)
end_date = datetime.date(2024, 12, 25)

# Создаем список рабочих дней
dates = pd.date_range(start=start_date, end=end_date, freq='B')  # 'B' - только будние дни

# Создаем пустой DataFrame для расписания
schedule_df = pd.DataFrame(index=students, columns=dates.date)

# Заполняем расписание: каждая строка соответствует ученику, и по кругу заполняем дежурства
for i, student in enumerate(students):
    assigned_dates = dates[i::len(students)]  # Распределяем даты через каждые n студентов
    for date in assigned_dates:
        schedule_df.loc[student, date.date()] = "X"

# Убираем год из столбцов (оставляем только день и месяц)
schedule_df.columns = schedule_df.columns.map(lambda d: d.strftime('%d-%m'))

# Сохраняем файл Excel с расписанием дежурств
schedule_output_filename = 'duty_schedule.xlsx'
with pd.ExcelWriter(schedule_output_filename, engine='openpyxl') as writer:
    schedule_df.to_excel(writer, sheet_name='Расписание дежурств')

    # Открываем книгу для закрашивания ячеек с дежурными
    workbook = writer.book
    sheet = workbook['Расписание дежурств']

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Желтый цвет

    # Закрашиваем ячейки, где указано "X"
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=len(students) + 1, max_col=len(dates) + 1):
        for cell in row:
            if cell.value == "X":
                cell.fill = fill

    # Устанавливаем форматирование для колонок с датами
    for col in sheet.columns:
        for cell in col:
            if cell.row == 1:
                cell.number_format = 'DD-MM'  # Формат даты без года

print(f"Файл '{schedule_output_filename}' успешно создан.")
